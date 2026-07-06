"""
build_region_map.py — Build site/campus -> region lookup from LaMP daily download.

Auto-picks the newest "*LaMP-daily-download.xlsx" from the Global Lab Planners
archive folder, reads the Spaces sheet (headers on row 7), and writes
site_region_map.json next to this script.

Output JSON shape:
{
  "generated_at": "...",
  "source": "<filename>",
  "building": { "AL4": { "region": "AMER", "site": "OR", "site_name": "Oregon, Hillsboro", "campus": "AL", "campus_name": "Aloha" }, ... },
  "campus":   { "AL":  { "region": "AMER", "site": "OR", "site_name": "Oregon, Hillsboro", "campus_name": "Aloha" }, ... },
  "site":     { "OR":  { "region": "AMER", "site_name": "Oregon, Hillsboro" }, ... }
}
"""
import json
import os
import sys
import glob
import shutil
import tempfile
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed. Run: py -m pip install openpyxl")
    sys.exit(1)

ARCHIVE_DIR = r"C:\Users\ylahat\OneDrive - Intel Corporation\Documents - Global Lab Planners\General\LaMP Download Archive"
OUTPUT      = os.path.join(os.path.dirname(__file__), "site_region_map.json")

# Spaces sheet layout (1-based column indices)
HEADER_ROW    = 7
COL_BUILDING  = 16   # P
COL_CAMPUS    = 17   # Q
COL_CAMPUS_NM = 18   # R
COL_REGION    = 22   # V
COL_REGION_NM = 23   # W
COL_SITE      = 24   # X
COL_SITE_NM   = 25   # Y

# Manual building -> mapping overrides.
# Use for legacy or missing-in-LaMP codes. Three formats:
#   1. "CODE": "campus:XX"  - alias to an existing LaMP campus
#   2. "CODE": "site:XX"    - alias to an existing LaMP site
#   3. "CODE": {"region": "...", "site": "...", "site_name": "..."}
#                          - explicit (used when LaMP doesn't have that region/site)
#
# Sources cross-checked against UDM (Yariv UDM.xlsx) and confirmed with user.
MANUAL_BUILDING_OVERRIDES = {
    # Gdansk legacy (first building never numbered)
    "IGK1":  "campus:IGK",

    # Hillsboro Oregon (Jones Farm campus)
    "FARM":  "campus:JF",

    # Haifa Israel — IDC = Matam, Haifa (UDM camp=HA); IDC7 already in LaMP
    "IDC":   "site:IS",
    "IDC10": "site:IS",

    # San Jose / Santa Clara California — bucket under SC site
    "SAN":   "site:SC",
    "SJI":   "site:SC",
    "SJI1":  "site:SC",
    "SJI2":  "site:SC",
    "SJI3":  "site:SC",
    "SJI4":  "site:SC",

    # Hyderabad India (CRESD: SKC = Salarpuria Sattva Knowledge City) — NOT Santa Clara
    "SKC":   {"region": "APAC", "site": "HY", "site_name": "India, Hyderabad"},
    "SKC1":  {"region": "APAC", "site": "HY", "site_name": "India, Hyderabad"},

    # Bengaluru India
    "BGA":   "site:BA",
    "BGA1":  "site:BA",
    "EMB":   "site:BA",   # CRESD: Embassy, Bangalore

    # Munich Germany (Lilienthalstrasse campus)
    "LIL1":  "site:MU",

    # New sites NOT in LaMP — introduce explicit region/site codes
    "TEC1":  {"region": "AMER", "site": "FC", "site_name": "Colorado, Fort Collins"},
    "ITK":   {"region": "APAC", "site": "TK", "site_name": "Japan, Tokyo"},
    "ITK1":  {"region": "APAC", "site": "TK", "site_name": "Japan, Tokyo"},
    "AMP1":  {"region": "AMER", "site": "PA", "site_name": "Pennsylvania, Allentown"},
    "ISW":   {"region": "EMEA", "site": "UK", "site_name": "United Kingdom, Swindon"},

    # Beijing China (CRESD: RYC=Raycom, UBP=Universal Business Park)
    "RYC":   "site:BJ",
    "UBP":   "site:BJ",
    "UBP1":  "site:BJ",

    # Seoul South Korea (CRESD: GLS = Glass Tower)
    "GLS":   "site:IK",

    # Taiwan (CRESD: GND = Gongdao — Hsinchu area)
    "GND":   "site:HS",
    "GND2":  "site:HS",

    # Tokyo / Yokohama Japan (CRESD: ISHI = Shinyokohama, Yokohama)
    "ISH":   "site:TK",

    # San Diego California (CRESD: SCS = Scripps, San Diego) — new site code SD
    "SCS":   {"region": "AMER", "site": "SD", "site_name": "California, San Diego"},
    "SCS1":  {"region": "AMER", "site": "SD", "site_name": "California, San Diego"},

    # Munich Germany — additional campus AMC (Am Campeon, Neubiberg)
    "AMC10": "site:MU",

    # Bengaluru India — additional campus ECO (Ecospace)
    "ECO1":  "site:BA",

    # Jerusalem Israel — campus JR (not a LaMP site, bucket under IS)
    "JER":   "site:IS",

    # Free-text fallbacks: SharePoint values that don't follow the building code pattern
    "FOLSOM":     "site:FM",
    "SANTA CLAR": "site:SC",
}


def find_latest_lamp():
    pattern = os.path.join(ARCHIVE_DIR, "*LaMP-daily-download.xlsx")
    files = glob.glob(pattern)
    if not files:
        raise FileNotFoundError(f"No LaMP files found in {ARCHIVE_DIR}")
    # Filename starts with YYYYMMDD so lexical sort matches date order
    files.sort(key=lambda p: os.path.basename(p), reverse=True)
    return files[0]


def main():
    src = find_latest_lamp()
    print(f"Source: {os.path.basename(src)}")

    # Copy to temp because OneDrive sometimes locks the original
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tf:
        tmp_path = tf.name
    shutil.copy2(src, tmp_path)

    try:
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        if "Spaces" not in wb.sheetnames:
            raise RuntimeError(f"'Spaces' sheet not found. Available: {wb.sheetnames}")
        ws = wb["Spaces"]

        building_map = {}
        campus_map   = {}
        site_map     = {}

        rows_processed = 0
        for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            building = row[COL_BUILDING - 1] if len(row) >= COL_BUILDING else None
            campus   = row[COL_CAMPUS   - 1] if len(row) >= COL_CAMPUS   else None
            camp_nm  = row[COL_CAMPUS_NM- 1] if len(row) >= COL_CAMPUS_NM else None
            region   = row[COL_REGION   - 1] if len(row) >= COL_REGION   else None
            reg_nm   = row[COL_REGION_NM- 1] if len(row) >= COL_REGION_NM else None
            site     = row[COL_SITE     - 1] if len(row) >= COL_SITE     else None
            site_nm  = row[COL_SITE_NM  - 1] if len(row) >= COL_SITE_NM  else None

            if not region or not site:
                continue
            rows_processed += 1

            region = str(region).strip().upper()
            site_s = str(site).strip().upper()
            site_nm_s = str(site_nm).strip() if site_nm else ""

            if site_s and site_s not in site_map:
                site_map[site_s] = {
                    "region":    region,
                    "region_name": str(reg_nm).strip() if reg_nm else "",
                    "site_name": site_nm_s,
                }

            if campus:
                camp_s = str(campus).strip().upper()
                if camp_s and camp_s not in campus_map:
                    campus_map[camp_s] = {
                        "region":      region,
                        "site":        site_s,
                        "site_name":   site_nm_s,
                        "campus_name": str(camp_nm).strip() if camp_nm else "",
                    }

            if building:
                bld_s = str(building).strip().upper()
                if bld_s and bld_s not in building_map:
                    building_map[bld_s] = {
                        "region":      region,
                        "site":        site_s,
                        "site_name":   site_nm_s,
                        "campus":      str(campus).strip().upper() if campus else "",
                        "campus_name": str(camp_nm).strip() if camp_nm else "",
                    }
        wb.close()
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    # Apply manual overrides (supports campus alias, site alias, or explicit dict)
    overrides_applied = 0
    for bld, spec in MANUAL_BUILDING_OVERRIDES.items():
        bld_u = bld.upper()
        if bld_u in building_map:
            continue  # already in LaMP, don't clobber

        info = None
        if isinstance(spec, dict):
            info = {
                "region":      spec["region"],
                "site":        spec.get("site", ""),
                "site_name":   spec.get("site_name", ""),
                "campus":      "",
                "campus_name": "",
            }
        elif isinstance(spec, str):
            if spec.startswith("campus:"):
                camp_u = spec[len("campus:"):].upper()
                ref = campus_map.get(camp_u)
                if not ref:
                    print(f"  WARNING: override {bld_u} -> {spec} skipped (campus not in LaMP)")
                    continue
                info = {
                    "region":      ref["region"],
                    "site":        ref["site"],
                    "site_name":   ref["site_name"],
                    "campus":      camp_u,
                    "campus_name": ref.get("campus_name", ""),
                }
            elif spec.startswith("site:"):
                site_u = spec[len("site:"):].upper()
                ref = site_map.get(site_u)
                if not ref:
                    print(f"  WARNING: override {bld_u} -> {spec} skipped (site not in LaMP)")
                    continue
                info = {
                    "region":      ref["region"],
                    "site":        site_u,
                    "site_name":   ref.get("site_name", ""),
                    "campus":      "",
                    "campus_name": "",
                }
            else:
                # legacy plain-string form: treat as campus alias
                ref = campus_map.get(spec.upper())
                if not ref:
                    print(f"  WARNING: override {bld_u} -> {spec} skipped (campus not in LaMP)")
                    continue
                info = {
                    "region":      ref["region"],
                    "site":        ref["site"],
                    "site_name":   ref["site_name"],
                    "campus":      spec.upper(),
                    "campus_name": ref.get("campus_name", ""),
                }
        if not info:
            continue
        info["override"] = True
        building_map[bld_u] = info

        # Also register any new site code so site filters know about it
        site_u = info["site"]
        if site_u and site_u not in site_map:
            site_map[site_u] = {
                "region":      info["region"],
                "region_name": "",
                "site_name":   info["site_name"],
            }
        overrides_applied += 1

    out = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "source":       os.path.basename(src),
        "overrides":    overrides_applied,
        "building":     building_map,
        "campus":       campus_map,
        "site":         site_map,
    }
    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"Rows processed: {rows_processed:,}")
    print(f"Buildings: {len(building_map):,}  |  Campuses: {len(campus_map):,}  |  Sites: {len(site_map):,}")
    print(f"Manual overrides applied: {overrides_applied}")
    print(f"Regions:   {sorted({v['region'] for v in site_map.values()})}")
    print(f"Sites:     {sorted(site_map.keys())}")
    print(f"Saved -> {OUTPUT}")


if __name__ == "__main__":
    main()
