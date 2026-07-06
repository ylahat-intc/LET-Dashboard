"""
LET Lab Space Request Dashboard - Excel Workbook Builder
Writes all 1,479+ records to "All Requests" as a proper Excel Table,
then uses Excel SUMIFS/COUNTIFS formulas in all analysis sheets so every
number is auditable and live-recalculates when data changes.

HOW TO USE:
  1. Run: py build_excel.py
  2. Open: LET_Dashboard.xlsx
  For live data: Data > Get Data > SharePoint List, then Refresh All
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timedelta
from collections import Counter
import json, os, re

# ── CONFIG ────────────────────────────────────────────────────────────────────
CACHE_FILE  = os.path.join(os.path.dirname(__file__), "let_data_cache.json")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "LET_Dashboard.xlsx")
COST_SQFT   = 692

# Colors
BLUE = "0068B5"; LT_BLUE = "D6EEFF"; DARK = "1A1A1A"
GREEN = "00A86B"; ORANGE = "FF8C00"; RED = "CC3333"
GRAY = "6B7280"; HEADER_BG = "0068B5"; HEADER_FG = "FFFFFF"
BAND1 = "EFF6FF"; BAND2 = "FFFFFF"; KPI_BG = "F0F7FF"
TOTAL_BG = "D6EEFF"; NOTE_BG = "FFFDE7"

def hf(h): return PatternFill(fill_type="solid", fgColor=h)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True, color=HEADER_FG, size=10)
        c.fill = hf(HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30

# ── DATA LOADING ──────────────────────────────────────────────────────────────
OLE_BASE = datetime(1899, 12, 30)

def parse_date(d):
    if not d: return None
    if isinstance(d, (int, float)) and d > 0:
        try: return OLE_BASE + timedelta(days=float(d))
        except: return None
    if isinstance(d, str):
        for fmt in ("%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S", "%m/%d/%Y", "%Y-%m-%d"):
            try: return datetime.strptime(d[:len(fmt)], fmt)
            except: continue
    return None

def parse_sqft(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return int(v)
    m = re.search(r'\d[\d,]*', str(v))
    return int(m.group().replace(",","")) if m else 0

def normalize_status(s):
    if not s: return "Unknown"
    sl = s.strip().lower()
    if "filled with existing" in sl: return "Reuse"
    if "return back" in sl: return "Lab Return"
    if "vetted, require cre" in sl or "lswg capital" in sl or "mrc approved" in sl: return "Capital/CRE"
    if "withdrawn" in sl: return "Withdrawn"
    if "denied" in sl or "unsupportable" in sl: return "Denied/Unsupportable"
    if "hold" in sl or "pending" in sl or "open" in sl or "new request" in sl: return "Open/Pending"
    return "Other"

BU_MAP = {
    "DCG (Data Center Group)": "DCG",
    "Silicon & Platform Engineering Group (SPE)": "SPE",
    "Client Computing Group-CCG": "CCG",
    "CTO&AI/IATG": "CTO&AI",
    "SIG VE grp": "SIG VE", "SIG MPE grp": "SIG MPE", "SIG WI grp": "SIG WI",
    "TBD - needs update": "TBD", "ICG (Intel China Group)": "ICG",
}

def clean_bu(bu):
    if not bu: return "Unknown"
    return BU_MAP.get(bu.strip(), bu.strip())

def load_rows():
    with open(CACHE_FILE, encoding="utf-8-sig") as f:
        raw = json.load(f)
    items = raw if isinstance(raw, list) else raw.get("items", [])
    rows = []
    for item in items:
        fld = item.get("fields", item)
        def g(*keys):
            for k in keys:
                v = fld.get(k)
                if v not in (None, "", 0, False): return v
            return None
        date_req   = parse_date(g("DateRequested","field_2"))
        date_close = parse_date(g("DateClosed","field_29","field_28"))
        sqft_req   = parse_sqft(g("SqFtRequested","field_13"))
        sqft_asn   = parse_sqft(g("SqFtAssigned","field_30"))
        planner    = str(g("Owner","Planner","field_26") or "").strip()
        if "/" in planner: planner = planner.split("/")[0].strip()
        days_open  = 0
        if date_req and date_close:
            d = (date_close - date_req).days
            if 0 <= d <= 1500: days_open = d
        rows.append({
            "id":        str(g("RequestID","Title") or ""),
            "status_r":  str(g("Status","field_3") or "").strip(),
            "status":    normalize_status(str(g("Status","field_3") or "")),
            "bu":        clean_bu(str(g("SuperGroup","field_7") or "")),
            "group":     str(g("Group","field_8") or "").strip(),
            "sqft_req":  sqft_req,
            "sqft_asn":  sqft_asn,
            "planner":   planner,
            "touch":     str(g("Touch","field_17") or "").strip(),
            "perm":      str(g("Permanent","field_9") or "").strip(),
            "cpa":       str(g("CPA","field_10") or "NO"),
            "benches":   parse_sqft(g("Benches","field_14")),
            "racks":     parse_sqft(g("Racks","field_15")),
            "title":     str(g("Title","field_27") or "")[:80],
            "site_to":   str(g("SiteTo","field_20") or ""),
            "site_from": str(g("SiteFrom","field_21") or ""),
            "date_req":  date_req,
            "date_cl":   date_close,
            "year":      date_req.year if date_req else None,
            "month":     date_req.strftime("%Y-%m") if date_req else "",
            "days_open": days_open,
        })
    return rows

# Column letters in "All Requests" sheet
# A=Request ID  B=Status(raw)  C=Status(Normalized)  D=BU  E=Group
# F=SqFt Req    G=SqFt Assigned  H=Planner  I=Touch  J=Permanent  K=CPA
# L=Benches  M=Racks  N=Date Requested  O=Date Closed  P=Year  Q=Month
# R=Title  S=Site To  T=Site From  U=Days Open
RAW = "'All Requests'"
C_STATUS  = "C"; C_BU      = "D"; C_SQFTREQ = "F"
C_SQFTASN = "G"; C_PLANNER = "H"; C_CPA     = "K"
C_YEAR    = "P"; C_DAYS    = "U"

def rng(col, last_row):
    """Absolute bounded range on All Requests sheet."""
    return f"{RAW}!${col}$2:${col}${last_row}"

# ── SHEET 1: All Requests ─────────────────────────────────────────────────────
def build_raw_data(wb, rows):
    ws = wb.active
    ws.title = "All Requests"
    ws.freeze_panes = "A2"

    cols   = ["Request ID","Status (Raw)","Status (Normalized)","BU","Group",
              "SqFt Requested","SqFt Assigned","Planner","Touch",
              "Permanent","CPA","Benches","Racks",
              "Date Requested","Date Closed","Year","Month",
              "Title","Site To","Site From","Days Open"]
    widths = [10,36,22,10,14,12,12,14,10,10,5,8,6,13,12,6,8,40,18,12,10]
    header_row(ws, 1, cols, widths)

    last_row = 1
    for i, r in enumerate(rows, 2):
        band = BAND1 if i % 2 == 0 else BAND2
        vals = [r["id"], r["status_r"], r["status"], r["bu"], r["group"],
                r["sqft_req"], r["sqft_asn"], r["planner"], r["touch"],
                r["perm"], r["cpa"], r["benches"], r["racks"],
                r["date_req"], r["date_cl"], r["year"], r["month"],
                r["title"], r["site_to"], r["site_from"], r["days_open"]]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.fill = hf(band)
            c.alignment = Alignment(vertical="center", wrap_text=(j == 18))
            if j in (6, 7, 12, 13): c.number_format = "#,##0"
            if j in (14, 15) and v:  c.number_format = "YYYY-MM-DD"
        last_row = i

    try:
        tbl = Table(displayName="tblRequests",
                    ref=f"A1:{get_column_letter(len(cols))}{last_row}")
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tbl)
    except Exception as e:
        print(f"  ⚠️ Table skipped: {e}")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    return last_row

# ── SHEET 2: Dashboard KPIs (Excel formulas) ──────────────────────────────────
def build_dashboard(wb, last_row):
    ws = wb.create_sheet("\U0001f4ca Dashboard")
    ws.sheet_view.showGridLines = False

    def R(col): return rng(col, last_row)

    # Banner
    ws.merge_cells("A1:P1")
    c = ws["A1"]
    c.value = "LET Lab Space Request \u2014 Management Dashboard"
    c.font = Font(bold=True, size=18, color=HEADER_FG)
    c.fill = hf(HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 42

    ws.merge_cells("A2:P2")
    c = ws["A2"]
    c.value = (f"Generated: {datetime.now():%Y-%m-%d %H:%M}  |  "
               f"Source: SharePoint LET Lab Space Request/Release  |  "
               f"Benchmark: ${COST_SQFT:,}/sqft new construction")
    c.font = Font(italic=True, size=10, color="666666")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Editable benchmark cell (C3) — all cost formulas reference this
    ws["A3"].value = "Cost Benchmark ($/sqft) \u2192"
    ws["A3"].font = Font(bold=True, size=10)
    ws.merge_cells("A3:B3")
    ws["C3"].value = COST_SQFT
    ws["C3"].number_format = '"$"#,##0'
    ws["C3"].font = Font(bold=True, size=12, color=GREEN)
    ws["C3"].fill = hf("E8F5E9")
    ws["C3"].alignment = Alignment(horizontal="center")
    ws.merge_cells("D3:P3")
    ws["D3"].value = "\u2190 Change this cell to update ALL cost calculations instantly"
    ws["D3"].font = Font(italic=True, size=9, color=GRAY)
    ws.row_dimensions[3].height = 22

    # KPI card rows
    ws.row_dimensions[4].height = 10
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 48
    ws.row_dimensions[7].height = 20
    ws.row_dimensions[8].height = 6

    kpis = [
        # (label, value_formula, sub_formula, number_format, color)
        ("\U0001f4b0 Cost Avoided",
         f'=SUMIF({R(C_STATUS)},"Reuse",{R(C_SQFTASN)})*$C$3',
         f'=TEXT(SUMIF({R(C_STATUS)},"Reuse",{R(C_SQFTASN)}),"#,##0")&" sqft \u00d7 $"&TEXT($C$3,"#,##0")',
         '"$"#,##0,,"M"', GREEN),

        ("\u267b\ufe0f Reuse Rate",
         f'=IFERROR(COUNTIF({R(C_STATUS)},"Reuse")/COUNTA({R("A")}),0)',
         f'=COUNTIF({R(C_STATUS)},"Reuse")&" reuse fills"',
         "0.0%", BLUE),

        ("\U0001f4d0 SqFt Reused",
         f'=SUMIF({R(C_STATUS)},"Reuse",{R(C_SQFTASN)})',
         '"Assigned sqft (official OKR)"',
         "#,##0", "0068B5"),

        ("\U0001f4e6 SqFt Returned",
         f'=SUMIF({R(C_STATUS)},"Lab Return",{R(C_SQFTREQ)})',
         '"Lab space returned to inventory"',
         "#,##0", ORANGE),

        ("\U0001f4cb Total Requests",
         f'=COUNTA({R("A")})',
         f'=COUNTIF({R(C_STATUS)},"Open/Pending")&" open now"',
         "#,##0", DARK),

        ("\u23f1\ufe0f Avg Days/Close",
         f'=IFERROR(AVERAGEIFS({R(C_DAYS)},{R(C_STATUS)},"<>Open/Pending",{R(C_DAYS)},">0"),0)',
         '"Days from submit to resolution"',
         "#,##0.0", GRAY),

        ("\U0001f534 CPA Requests",
         f'=COUNTIF({R(C_CPA)},"YES")+COUNTIF({R(C_CPA)},"Yes")',
         '"Critical Path flagged"',
         "#,##0", RED),

        ("\U0001f3d7\ufe0f Capital/CRE",
         f'=COUNTIF({R(C_STATUS)},"Capital/CRE")',
         '"Escalated to CRE/Capital"',
         "#,##0", ORANGE),
    ]

    for idx, (label, val_f, sub_f, fmt, color) in enumerate(kpis):
        col = idx * 2 + 1
        ws.column_dimensions[get_column_letter(col)].width   = 11
        ws.column_dimensions[get_column_letter(col+1)].width = 11
        for r in (5, 6, 7, 8):
            ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+1)

        c = ws.cell(row=5, column=col, value=label)
        c.font = Font(bold=True, size=10, color=HEADER_FG); c.fill = hf(HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = ws.cell(row=6, column=col, value=val_f)
        c.font = Font(bold=True, size=20, color=color); c.fill = hf(KPI_BG)
        c.number_format = fmt; c.alignment = Alignment(horizontal="center", vertical="center")

        c = ws.cell(row=7, column=col, value=sub_f)
        c.font = Font(italic=True, size=9, color=GRAY); c.fill = hf(KPI_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=8, column=col).fill = hf(color)

    # Yearly snapshot table (rows 10+)
    ws.row_dimensions[9].height = 10
    ws.merge_cells("A10:H10")
    ws["A10"].value = "\U0001f4c5 Yearly Snapshot (all categories)"
    ws["A10"].font = Font(bold=True, size=12, color=BLUE)

    header_row(ws, 11,
        ["Year","# Requests","Reuse Fills","Open Now","SqFt Assigned","Cost Avoided ($)","Reuse Rate","Avg Days"],
        [8,11,11,10,14,16,11,11])

    for i, yr in enumerate(range(2015, 2031), 12):
        row = i
        band = BAND1 if i % 2 == 0 else BAND2
        ws.cell(row=row, column=1, value=yr).fill = hf(band)
        ws.cell(row=row, column=1).number_format = "0"
        a_yr = f"$A${row}"
        formulas = [
            f'=COUNTIFS({R(C_YEAR)},{a_yr})',
            f'=COUNTIFS({R(C_STATUS)},"Reuse",{R(C_YEAR)},{a_yr})',
            f'=COUNTIFS({R(C_STATUS)},"Open/Pending",{R(C_YEAR)},{a_yr})',
            f'=SUMIFS({R(C_SQFTASN)},{R(C_STATUS)},"Reuse",{R(C_YEAR)},{a_yr})',
            f'=E{row}*$C$3',
            f'=IFERROR(C{row}/B{row},0)',
            f'=IFERROR(AVERAGEIFS({R(C_DAYS)},{R(C_YEAR)},{a_yr},{R(C_DAYS)},">0"),0)',
        ]
        fmts = ["#,##0","#,##0","#,##0","#,##0",'"$"#,##0,,"M"',"0.0%","#,##0.0"]
        for j, (f2, fm) in enumerate(zip(formulas, fmts), 2):
            c = ws.cell(row=row, column=j, value=f2)
            c.fill = hf(band); c.number_format = fm
            c.alignment = Alignment(horizontal="right")

    trow = 12 + 16
    for col in range(1, 9): ws.cell(row=trow, column=col).fill = hf(TOTAL_BG)
    ws.cell(row=trow, column=1, value="TOTAL").font = Font(bold=True)
    for col_idx, fm in [(2,"#,##0"),(3,"#,##0"),(5,"#,##0"),(6,'"$"#,##0,,"M"')]:
        c = ws.cell(row=trow, column=col_idx,
                    value=f"=SUM({get_column_letter(col_idx)}12:{get_column_letter(col_idx)}{trow-1})")
        c.font = Font(bold=True); c.number_format = fm

    # Note
    nr = trow + 2
    ws.merge_cells(f"A{nr}:P{nr}")
    n = ws.cell(row=nr, column=1,
        value=("\u26a0\ufe0f  Cost Avoided = SqFt Assigned (col G in All Requests) \u00d7 benchmark. "
               "Pre-2022 records may have SqFt Assigned = 0 (not recorded). "
               "Official OKR ~$484M cumulative as of end 2025 uses assigned sqft only."))
    n.font = Font(italic=True, size=9, color="AA6600")
    n.fill = hf(NOTE_BG); n.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[nr].height = 28

    return ws

# ── SHEET 3: Cost Avoidance (formulas) ────────────────────────────────────────
def build_cost_avoidance(wb, last_row):
    ws = wb.create_sheet("\U0001f4b0 Cost Avoidance")
    ws.sheet_view.showGridLines = False

    def R(col): return rng(col, last_row)

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "Cost Avoidance Analysis \u2014 Lab Reuse vs New Build Construction"
    c.font = Font(bold=True, size=16, color=HEADER_FG)
    c.fill = hf(HEADER_BG); c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws["A2"].value = "Benchmark cost/sqft (editable):"
    ws["A2"].font = Font(bold=True, size=10)
    ws.merge_cells("A2:B2")
    ws["C2"].value = COST_SQFT
    ws["C2"].number_format = '"$"#,##0'
    ws["C2"].font = Font(bold=True, color=GREEN, size=12)
    ws["C2"].fill = hf("E8F5E9")
    ws.merge_cells("D2:I2")
    ws["D2"].value = "\u2190 Change to update all cost figures. Intel standard: $692/sqft new lab."
    ws["D2"].font = Font(italic=True, size=9, color=GRAY)
    ws.row_dimensions[2].height = 22

    for i, w in enumerate([8,16,14,14,16,14,14,14,14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A4:I4")
    ws["A4"].value = "\U0001f4c5 By Year"
    ws["A4"].font = Font(bold=True, size=12, color=BLUE)
    ws.row_dimensions[4].height = 24

    header_row(ws, 5,
        ["Year","# Reuse Fills","SqFt Requested","SqFt Assigned","Cost Avoided",
         "% of Total","Benchmark/sqft","# Closed","Reuse Rate"])

    for i, yr in enumerate(range(2015, 2031), 6):
        row = i
        band = BAND1 if i % 2 == 0 else BAND2
        ws.cell(row=row, column=1, value=yr).fill = hf(band)
        ws.cell(row=row, column=1).number_format = "0"
        a_yr = f"$A${row}"
        fs = [
            f'=COUNTIFS({R(C_STATUS)},"Reuse",{R(C_YEAR)},{a_yr})',
            f'=SUMIFS({R(C_SQFTREQ)},{R(C_STATUS)},"Reuse",{R(C_YEAR)},{a_yr})',
            f'=SUMIFS({R(C_SQFTASN)},{R(C_STATUS)},"Reuse",{R(C_YEAR)},{a_yr})',
            f'=D{row}*$C$2',
            f'=IFERROR(E{row}/SUM($E$6:$E$21),0)',
            f'=$C$2',
            f'=COUNTIFS({R(C_STATUS)},"<>Open/Pending",{R(C_STATUS)},"<>Unknown",{R(C_YEAR)},{a_yr})',
            f'=IFERROR(B{row}/H{row},0)',
        ]
        fmts = ["#,##0","#,##0","#,##0",'"$"#,##0',"0.0%",'"$"#,##0',"#,##0","0.0%"]
        for j, (f2, fm) in enumerate(zip(fs, fmts), 2):
            c = ws.cell(row=row, column=j, value=f2)
            c.fill = hf(band); c.number_format = fm
            c.alignment = Alignment(horizontal="right")
            if j == 5: c.font = Font(bold=True, color=GREEN)

    total_row = 22
    for col in range(1, 10): ws.cell(row=total_row, column=col).fill = hf(TOTAL_BG)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for col_idx, fm in [(2,"#,##0"),(3,"#,##0"),(4,"#,##0"),(5,'"$"#,##0')]:
        c = ws.cell(row=total_row, column=col_idx,
                    value=f"=SUM({get_column_letter(col_idx)}6:{get_column_letter(col_idx)}21)")
        c.font = Font(bold=True, size=12, color=(GREEN if col_idx == 5 else DARK))
        c.number_format = fm

    # Chart
    chart = BarChart()
    chart.type = "col"; chart.title = "Cost Avoided by Year ($)"
    chart.y_axis.title = "Dollars ($)"; chart.x_axis.title = "Year"
    data_ref = Reference(ws, min_col=5, max_col=5, min_row=5, max_row=21)
    cats_ref  = Reference(ws, min_col=1, max_col=1, min_row=6, max_row=21)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 24; chart.height = 14
    ws.add_chart(chart, "A25")

    # By status table
    s_start = 42
    ws.merge_cells(f"A{s_start}:I{s_start}")
    ws.cell(row=s_start, column=1, value="\U0001f4ca By Status Category").font = Font(bold=True, size=12, color=BLUE)
    header_row(ws, s_start+1, ["Status Category","Count","SqFt Requested","SqFt Assigned","Cost Avoided"])
    statuses = ["Reuse","Lab Return","Capital/CRE","Withdrawn","Denied/Unsupportable","Open/Pending","Other"]
    for i, st in enumerate(statuses, s_start+2):
        band = BAND1 if i % 2 == 0 else BAND2
        ws.cell(row=i, column=1, value=st).fill = hf(band)
        fs2 = [
            f'=COUNTIF({R(C_STATUS)},"{st}")',
            f'=SUMIF({R(C_STATUS)},"{st}",{R(C_SQFTREQ)})',
            f'=SUMIF({R(C_STATUS)},"{st}",{R(C_SQFTASN)})',
            f'=IF("{st}"="Reuse",D{i}*$C$2,"")',
        ]
        for j, (f2, fm) in enumerate(zip(fs2, ["#,##0","#,##0","#,##0",'"$"#,##0']), 2):
            c = ws.cell(row=i, column=j, value=f2)
            c.fill = hf(band); c.number_format = fm
            c.alignment = Alignment(horizontal="right")
    return ws

# ── SHEET 4: BU Analysis (formulas) ───────────────────────────────────────────
def build_bu_analysis(wb, rows, last_row):
    ws = wb.create_sheet("\U0001f3e2 BU Analysis")
    ws.sheet_view.showGridLines = False

    def R(col): return rng(col, last_row)

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "Business Unit Analysis"
    c.font = Font(bold=True, size=16, color=HEADER_FG)
    c.fill = hf(HEADER_BG); c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    for i, w in enumerate([20,11,11,11,14,11,13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A3:G3")
    ws["A3"].value = "\U0001f3e2 By Business Unit (COUNTIFS/SUMIFS formulas)"
    ws["A3"].font = Font(bold=True, size=12, color=BLUE)
    header_row(ws, 4, ["Business Unit","Total Reqs","Reuse Fills","Open Now","SqFt Requested","Reuse Rate","Capital/CRE"])

    bu_count = Counter(r["bu"] for r in rows if r["bu"] not in ("Unknown","TBD",""))
    for i, (bu, _) in enumerate(bu_count.most_common(20), 5):
        band = BAND1 if i % 2 == 0 else BAND2
        ws.cell(row=i, column=1, value=bu).fill = hf(band)
        a_bu = f"$A${i}"
        fs = [
            f'=COUNTIF({R(C_BU)},{a_bu})',
            f'=COUNTIFS({R(C_BU)},{a_bu},{R(C_STATUS)},"Reuse")',
            f'=COUNTIFS({R(C_BU)},{a_bu},{R(C_STATUS)},"Open/Pending")',
            f'=SUMIF({R(C_BU)},{a_bu},{R(C_SQFTREQ)})',
            f'=IFERROR(C{i}/B{i},0)',
            f'=COUNTIFS({R(C_BU)},{a_bu},{R(C_STATUS)},"Capital/CRE")',
        ]
        for j, (f2, fm) in enumerate(zip(fs, ["#,##0","#,##0","#,##0","#,##0","0.0%","#,##0"]), 2):
            c = ws.cell(row=i, column=j, value=f2)
            c.fill = hf(band); c.number_format = fm
            c.alignment = Alignment(horizontal="right")
    return ws

# ── SHEET 5: Planner Workload (formulas) ──────────────────────────────────────
def build_planner_sheet(wb, rows, last_row):
    ws = wb.create_sheet("\U0001f464 Planner Workload")
    ws.sheet_view.showGridLines = False

    def R(col): return rng(col, last_row)

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Planner Workload & Performance"
    c.font = Font(bold=True, size=16, color=HEADER_FG)
    c.fill = hf(HEADER_BG); c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    for i, w in enumerate([22,12,12,12,12,12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A3:F3")
    ws["A3"].value = "\U0001f464 By Planner (COUNTIFS/AVERAGEIFS formulas)"
    ws["A3"].font = Font(bold=True, size=12, color=BLUE)
    header_row(ws, 4, ["Planner","Total Handled","Reuse Closes","Open Now","Reuse Rate","Avg Days"])

    p_count = Counter(r["planner"] for r in rows if r["planner"])
    for i, (pl, _) in enumerate(p_count.most_common(15), 5):
        band = BAND1 if i % 2 == 0 else BAND2
        ws.cell(row=i, column=1, value=pl).fill = hf(band)
        a_pl = f"$A${i}"
        fs = [
            f'=COUNTIF({R(C_PLANNER)},{a_pl})',
            f'=COUNTIFS({R(C_PLANNER)},{a_pl},{R(C_STATUS)},"Reuse")',
            f'=COUNTIFS({R(C_PLANNER)},{a_pl},{R(C_STATUS)},"Open/Pending")',
            f'=IFERROR(C{i}/B{i},0)',
            f'=IFERROR(AVERAGEIFS({R(C_DAYS)},{R(C_PLANNER)},{a_pl},{R(C_DAYS)},">0"),0)',
        ]
        for j, (f2, fm) in enumerate(zip(fs, ["#,##0","#,##0","#,##0","0.0%","#,##0.0"]), 2):
            c = ws.cell(row=i, column=j, value=f2)
            c.fill = hf(band); c.number_format = fm
            c.alignment = Alignment(horizontal="right")
    return ws

# ── SHEET 6: README ───────────────────────────────────────────────────────────
def build_readme(wb):
    ws = wb.create_sheet("\u2139\ufe0f README")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100

    lines = [
        ("LET Lab Space Dashboard \u2014 Excel Guide", 16, HEADER_BG, HEADER_FG, True),
        ("", 11, "FFFFFF", "000000", False),
        ("SHEETS IN THIS WORKBOOK", 12, LT_BLUE, DARK, True),
        ("Dashboard  \u2014 KPI summary. ALL values are live Excel formulas. Change $C$3 to update cost.", 11, BAND1, "000000", False),
        ("Cost Avoidance  \u2014 Year-by-year + by-status. SUMIFS formulas. Change $C$2 for benchmark.", 11, BAND2, "000000", False),
        ("BU Analysis  \u2014 COUNTIFS/SUMIFS by business unit.", 11, BAND1, "000000", False),
        ("Planner Workload  \u2014 COUNTIFS/AVERAGEIFS by planner.", 11, BAND2, "000000", False),
        ("All Requests  \u2014 Full raw data as Excel Table 'tblRequests'. Filter/sort here.", 11, BAND1, "000000", False),
        ("", 11, "FFFFFF", "000000", False),
        ("HOW TO REFRESH DATA", 12, LT_BLUE, DARK, True),
        ("Option A: Double-click Refresh_Dashboard.bat  (runs Python pipeline, rebuilds this file)", 11, BAND1, "000000", False),
        ("Option B: Data > Get Data > SharePoint List > https://intel.sharepoint.com/sites/GlobalLabsandDataCenters", 11, BAND2, "000000", False),
        ("         Select 'LET Lab Space Request/Release' > Load. Then Refresh All updates everything.", 11, BAND1, "000000", False),
        ("", 11, "FFFFFF", "000000", False),
        ("COST AVOIDANCE METHODOLOGY", 12, LT_BLUE, DARK, True),
        (f"Benchmark: ${COST_SQFT}/sqft = Intel internal cost to build new lab space", 11, NOTE_BG, "AA6600", False),
        ("Formula: SqFt Assigned (col G) x benchmark, for all 'Reuse' status records", 11, NOTE_BG, "AA6600", False),
        ("Note: Pre-2022 records often have SqFt Assigned = 0 (not entered). Official OKR ~$484M (end 2025).", 11, NOTE_BG, "AA6600", False),
        ("SqFt Requested gives higher number; official uses assigned sqft only.", 11, NOTE_BG, "AA6600", False),
        ("", 11, "FFFFFF", "000000", False),
        ("POWER BI SETUP", 12, LT_BLUE, DARK, True),
        ("1. Power BI Desktop > Get Data > SharePoint Online List", 11, BAND1, "000000", False),
        ("2. Site: https://intel.sharepoint.com/sites/GlobalLabsandDataCenters", 11, BAND2, "000000", False),
        ("3. Select: LET Lab Space Request/Release", 11, BAND1, "000000", False),
        ("4. Build visuals with same KPIs as this dashboard", 11, BAND2, "000000", False),
        ("5. Publish to Power BI Service > Embed in SharePoint page for live dashboard", 11, BAND1, "000000", False),
    ]

    for i, (text, size, bg, fg, bold) in enumerate(lines, 1):
        ws.merge_cells(f"A{i}:B{i}")
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=size, color=fg)
        c.fill = hf(bg)
        c.alignment = Alignment(vertical="center", indent=2 if not bold else 0)
        ws.row_dimensions[i].height = 22 if text else 8

# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print("Building LET Dashboard Excel (formula-based)...")
    rows = load_rows()
    print(f"  Loaded {len(rows)} records")

    wb = openpyxl.Workbook()

    last_row = build_raw_data(wb, rows)
    print(f"  All Requests: {last_row-1} data rows + Excel Table registered")

    build_dashboard(wb, last_row)
    print("  Dashboard: all KPIs use SUMIFS/COUNTIFS Excel formulas")

    build_cost_avoidance(wb, last_row)
    print("  Cost Avoidance: SUMIFS by year + by status")

    build_bu_analysis(wb, rows, last_row)
    print("  BU Analysis: COUNTIFS/SUMIFS by business unit")

    build_planner_sheet(wb, rows, last_row)
    print("  Planner Workload: COUNTIFS/AVERAGEIFS by planner")

    build_readme(wb)

    wb.active = wb["\U0001f4ca Dashboard"]
    wb.save(OUTPUT_FILE)
    print(f"  Saved: {OUTPUT_FILE}")
    print()
    print("  Open in Excel. Every number is a live formula.")
    print("  Change $C$3 on Dashboard (or $C$2 on Cost Avoidance) to update ALL cost figures.")

if __name__ == "__main__":
    main()
